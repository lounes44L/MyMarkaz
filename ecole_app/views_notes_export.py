from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Avg, Max, Min
from .models import NoteExamen, Classe, Eleve
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

@login_required
def export_notes_excel(request):
    """Exporte les notes des élèves en Excel"""
    
    # Récupération de la composante
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder aux notes.")
        return redirect('selection_composante')
    
    # Vérifier si l'utilisateur est un professeur ou un administrateur
    is_admin = request.user.is_superuser or request.user.is_staff
    is_professeur = hasattr(request.user, 'professeur')
    
    if not (is_admin or is_professeur):
        messages.error(request, "Accès réservé aux professeurs et administrateurs.")
        return redirect('dashboard')
    
    # Filtrage par classe
    classe_id = request.GET.get('classe')
    selected_classe = None
    if classe_id:
        try:
            if is_professeur:
                # Si c'est un professeur, vérifier que la classe lui appartient
                selected_classe = Classe.objects.get(
                    id=classe_id,
                    professeur=request.user.professeur,
                    composante_id=composante_id
                )
            else:
                # Pour les administrateurs, pas de restriction
                selected_classe = Classe.objects.get(id=classe_id, composante_id=composante_id)
        except Classe.DoesNotExist:
            pass
    
    # Filtrage par élève
    eleve_id = request.GET.get('eleve')
    selected_eleve = None
    if eleve_id:
        try:
            selected_eleve = Eleve.objects.get(id=eleve_id, composante_id=composante_id)
        except Eleve.DoesNotExist:
            pass
    
    # Récupérer les notes
    if is_admin:
        notes_query = NoteExamen.objects.filter(classe__composante_id=composante_id).select_related('eleve', 'classe', 'professeur')
    else:
        professeur = request.user.professeur
        notes_query = NoteExamen.objects.filter(
            professeur=professeur, 
            classe__composante_id=composante_id
        ).select_related('eleve', 'classe')
    
    # Appliquer les filtres supplémentaires
    if selected_classe:
        notes_query = notes_query.filter(classe=selected_classe)
    
    if selected_eleve:
        notes_query = notes_query.filter(eleve=selected_eleve)
    
    # Trier les résultats
    notes = notes_query.order_by('classe__nom', 'eleve__nom', 'eleve__prenom', '-date_examen')
    
    # Statistiques globales
    stats = {
        'total': notes.count(),
        'moyenne': notes.aggregate(Avg('note'))['note__avg'] or 0,
        'max': notes.aggregate(Max('note'))['note__max'] or 0,
        'min': notes.aggregate(Min('note'))['note__min'] or 0,
    }
    
    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notes {selected_classe.nom if selected_classe else 'Toutes les classes'}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        'Nom',
        'Prénom',
        'Classe',
        'Titre',
        'Type d\'examen',
        'Note',
        'Note max',
        'Pourcentage',
        'Date examen',
        'Professeur',
        'Commentaire'
    ]
    
    # Ajout des en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Ajout des données
    row_num = 2
    for note in notes:
        # Calcul du pourcentage
        pourcentage = 0
        if note.note_max and note.note_max > 0:
            pourcentage = (note.note / note.note_max) * 100
        
        data_row = [
            note.eleve.nom,
            note.eleve.prenom,
            note.classe.nom,
            note.titre,
            note.get_type_examen_display(),
            note.note,
            note.note_max,
            f"{pourcentage:.1f}%",
            note.date_examen.strftime('%d/%m/%Y') if note.date_examen else '',
            note.professeur.nom if note.professeur else '',
            note.commentaire or ''
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top")
            cell.border = border
        
        row_num += 1
    
    # Ajout d'une ligne pour les statistiques globales
    row_num += 1
    ws.cell(row=row_num, column=1, value="Statistiques globales").font = Font(bold=True)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    
    row_num += 1
    stats_headers = ["Nombre de notes", "Moyenne", "Note maximale", "Note minimale"]
    stats_values = [stats['total'], f"{stats['moyenne']:.2f}", stats['max'], stats['min']]
    
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
    
    row_num += 1
    for col, value in enumerate(stats_values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = border
    
    # Ajustement des largeurs de colonnes
    column_widths = [15, 15, 12, 25, 15, 8, 8, 10, 12, 20, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Création de la réponse HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Nom du fichier
    today = datetime.datetime.now().strftime("%Y%m%d")
    filename = f"notes_{selected_classe.nom if selected_classe else 'toutes_classes'}_{today}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

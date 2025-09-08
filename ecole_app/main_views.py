from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models import Q, Count, Avg, F, Sum
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.utils.safestring import mark_safe
from django.core.mail import send_mail
from django.conf import settings
import csv
import json
import random
import string
from datetime import datetime
import datetime
from .models import Eleve, Professeur, Classe, Creneau, Paiement, ListeAttente, generer_mot_de_passe
from .forms import EleveForm, EleveRapideForm, ProfesseurForm, ClasseForm, CreneauForm, PaiementForm, ImportDataForm, ExportDataForm, DesarchivageEleveForm, ListeAttenteForm
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

# Fonctions utilitaires pour la génération d'identifiants et mots de passe
def generate_username(nom, prenom):
    """Génère un nom d'utilisateur au format prenom.nom"""
    # Supprime les espaces et caractères spéciaux
    nom_clean = ''.join(c for c in nom.lower() if c.isalnum())
    prenom_clean = ''.join(c for c in prenom.lower() if c.isalnum())
    
    # Créer l'identifiant au format prenom.nom
    return f"{prenom_clean}.{nom_clean}"

def generate_password(nom_complet):
    """Génère un mot de passe aléatoire basé sur le nom complet"""
    import random
    import string
    
    # Extraire le nom et prénom du nom complet si possible
    parts = nom_complet.split() if nom_complet else []
    
    # Utiliser les 3 premières lettres du nom et prénom si disponibles
    prefix = ""
    if parts and len(parts) > 0:
        prefix += parts[0][:3].lower()  # Première partie (nom)
    if parts and len(parts) > 1:
        prefix += parts[1][:3].lower()  # Deuxième partie (prénom)
    
    # Si le préfixe est trop court, ajouter des caractères aléatoires
    while len(prefix) < 4:
        prefix += random.choice(string.ascii_lowercase)
    
    # Ajouter 4 chiffres aléatoires
    suffix = ''.join(random.choices(string.digits, k=4))
    
    return prefix + suffix

# Vue principale - Dashboard
@login_required
def dashboard(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder au dashboard.")
        return redirect('selection_composante')
        
    # Statistiques générales - filtrées par composante
    total_eleves = Eleve.objects.filter(archive=False, composante_id=composante_id).count()
    total_profs = Professeur.objects.filter(composantes__id=composante_id).count()
    total_classes = Classe.objects.filter(composante_id=composante_id).count()
    
    # Calcul des statistiques d'occupation des classes - filtrées par composante
    classes = Classe.objects.filter(composante_id=composante_id)
    classe_stats = []
    total_capacite = 0
    
    for classe in classes:
        eleves_en_classe = classe.eleves.filter(archive=False).count()
        capacite = classe.capacite or 20
        total_capacite += capacite
        taux_occupation = min(100, round((eleves_en_classe / capacite) * 100)) if capacite > 0 else 0
        
        classe_stats.append({
            'id': classe.id,
            'nom': classe.nom,
            'eleves_en_classe': eleves_en_classe,
            'capacite': capacite,
            'taux_occupation': taux_occupation,
            'creneau': classe.creneau,
            'professeur': classe.professeur,
        })
    
    # Taux de remplissage global
    taux_remplissage = round((total_eleves / total_capacite) * 100) if total_capacite > 0 else 0
    
    # Statistiques financières
    total_attendu = Eleve.objects.aggregate(total=Sum('montant_total'))['total'] or 0
    total_collecte = Paiement.objects.aggregate(total=Sum('montant'))['total'] or 0
    pourcentage_collecte = round((total_collecte / total_attendu) * 100) if total_attendu > 0 else 0

    context = {
        'total_eleves': total_eleves,
        'total_profs': total_profs,
        'total_classes': total_classes,
        'total_creneaux': Creneau.objects.filter(composante_id=composante_id).count(),
        'classe_stats': classe_stats,
        'taux_remplissage': taux_remplissage,
        'total_capacite': total_capacite,
        'total_attendu': total_attendu,
        'total_collecte': total_collecte,
        'pourcentage_collecte': pourcentage_collecte,
    }

    return render(request, 'ecole_app/dashboard.html', context)

# Gestion des élèves

@login_required
def liste_attente(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la liste d'attente.")
        return redirect('selection_composante')
        
    attente_tab = True
    eleves_attente = ListeAttente.objects.filter(ajoute_definitivement=False, composante_id=composante_id).order_by('-date_ajout')
    form = ListeAttenteForm(request.POST or None)
    if request.method == 'POST' and 'ajout_attente' in request.POST:
        if form.is_valid():
            # Ajouter la composante avant de sauvegarder
            eleve_attente = form.save(commit=False)
            eleve_attente.composante_id = composante_id
            eleve_attente.save()
            messages.success(request, "L'élève a été ajouté à la liste d'attente.")
            return redirect('liste_attente')
    # Pour le modal d'ajout définitif : on fournit la liste des classes et créneaux de la composante actuelle
    classes = Classe.objects.filter(composante_id=composante_id)
    creneaux = Creneau.objects.filter(composante_id=composante_id)
    return render(request, 'ecole_app/eleves/liste_attente.html', {
        'attente_tab': True,
        'eleves_attente': eleves_attente,
        'form': form,
        'classes': classes,
        'creneaux': creneaux,
    })

@login_required
def modifier_eleve_attente(request, eleve_id):
    eleve = get_object_or_404(ListeAttente, id=eleve_id, ajoute_definitivement=False)
    if request.method == 'POST':
        form = ListeAttenteForm(request.POST, instance=eleve)
        if form.is_valid():
            form.save()
            messages.success(request, "Les informations de l'élève ont été modifiées.")
            return redirect('liste_attente')
    else:
        form = ListeAttenteForm(instance=eleve)
    # On réutilise le template d'ajout mais en mode édition
    return render(request, 'ecole_app/eleves/modifier_eleve_attente.html', {
        'form': form,
        'eleve': eleve
    })

@login_required
def supprimer_eleve_attente(request, eleve_id):
    eleve = get_object_or_404(ListeAttente, id=eleve_id, ajoute_definitivement=False)
    if request.method == 'POST':
        eleve.delete()
        messages.success(request, "L'élève a été supprimé de la liste d'attente.")
    return redirect('liste_attente')

@login_required
def ajouter_definitivement(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour ajouter un élève.")
        return redirect('selection_composante')
        
    if request.method == 'POST':
        enfant_id = request.POST.get('enfant_id')
        if not enfant_id:
            messages.error(request, "Données invalides.")
            return redirect('liste_attente')
        try:
            enfant = ListeAttente.objects.get(id=enfant_id, ajoute_definitivement=False, composante_id=composante_id)
        except ListeAttente.DoesNotExist:
            messages.error(request, "Données invalides.")
            return redirect('liste_attente')
        # Créer un nouvel élève à partir des infos de l'enfant
        eleve = Eleve.objects.create(
            nom=enfant.nom,
            prenom=enfant.prenom,
            date_naissance=enfant.date_naissance,
            telephone=enfant.telephone,
            email=enfant.email,
            remarque=enfant.remarque,
            composante_id=composante_id,
            archive=False
        )
        enfant.ajoute_definitivement = True
        enfant.save()
        messages.success(request, f"{eleve.prenom} {eleve.nom} a été ajouté comme élève actif.")
        return redirect('liste_attente')
    return redirect('liste_attente')

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
import pandas as pd
from openpyxl import load_workbook
import datetime
import io

@login_required
def eleve_home(request):
    eleve = getattr(request.user, 'eleve', None)
    if not eleve:
        return redirect('dashboard')
    
    # Récupérer les notes d'examen de l'élève avec optimisation des requêtes
    notes_examens = eleve.notes_examens.select_related('professeur', 'classe').order_by('-date_examen')[:10]
    
    # Calculer les statistiques des notes
    if notes_examens:
        notes_list = [note.pourcentage for note in notes_examens]
        moyenne_generale = round(sum(notes_list) / len(notes_list), 1)
        derniere_note = notes_examens[0] if notes_examens else None
    else:
        moyenne_generale = 0
        derniere_note = None
    
    context = {
        'eleve': eleve,
        'notes_examens': notes_examens,
        'moyenne_generale': moyenne_generale,
        'derniere_note': derniere_note,
        'nb_notes': notes_examens.count() if notes_examens else 0
    }
    
    return render(request, 'ecole_app/eleves/home.html', context)

from django.views.decorators.http import require_POST

@require_POST
@login_required
def archiver_eleve(request):
    eleve_id = request.POST.get('eleve_id')
    motif = request.POST.get('motif_archive', '').strip()
    eleve = get_object_or_404(Eleve, pk=eleve_id, archive=False)
    if motif:
        eleve.archive = True
        eleve.motif_archive = motif
        eleve.save()
        messages.success(request, f"L'élève {eleve.nom} {eleve.prenom} a été archivé.")
    else:
        messages.error(request, "Le motif d'archivage est obligatoire.")
    return redirect('liste_eleves')

@require_POST
@login_required
def desarchiver_eleve(request):
    eleve_id = request.POST.get('eleve_id')
    eleve = get_object_or_404(Eleve, pk=eleve_id, archive=True)
    # Désarchiver directement sans formulaire
    eleve.archive = False
    eleve.motif_archive = ''
    eleve.save()
    messages.success(request, f"L'élève {eleve.nom} {eleve.prenom} a été désarchivé.")
    return redirect('archives_eleves')

@login_required
def archives_eleves(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder aux archives.")
        return redirect('selection_composante')
        
    # Filtres pour le tri
    classe_id = request.GET.get('classe')
    creneau_id = request.GET.get('creneau')
    eleves = Eleve.objects.filter(archive=True, composante_id=composante_id)
    if classe_id:
        eleves = eleves.filter(classes__id=classe_id)
    if creneau_id:
        eleves = eleves.filter(creneau_id=creneau_id)
    eleves = eleves.order_by('nom', 'prenom')
    form = EleveForm()
    form_rapide = EleveRapideForm()
    from .models import Classe, Creneau
    classes = Classe.objects.filter(composante_id=composante_id).order_by('nom')
    creneaux = Creneau.objects.filter(composante_id=composante_id).order_by('nom')
    context = {
        'eleves': eleves,
        'form': form,
        'form_rapide': form_rapide,
        'archives_tab': True,
        'classes': classes,
        'creneaux': creneaux,
    }
    return render(request, 'ecole_app/eleves/liste.html', context)

@login_required
def liste_eleves(request):
    # Import des modèles nécessaires
    from .models import Classe, Creneau
    
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la liste des élèves.")
        return redirect('selection_composante')
        
    # Filtres pour le tri
    classe_id = request.GET.get('classe')
    creneau_id = request.GET.get('creneau')
    
    # Construction de la requête de base
    query = Q(composante_id=composante_id) & Q(archive=False)
    
    # Ajouter les filtres si nécessaire
    if classe_id:
        query &= Q(classes__id=classe_id)
    if creneau_id:
        query &= Q(creneau__id=creneau_id)
    
    # Exécuter la requête finale
    eleves = Eleve.objects.filter(query).order_by('nom', 'prenom').distinct()
    form = EleveForm(request=request)  # Utiliser le formulaire complet par défaut
    form_rapide = EleveRapideForm(request=request)  # Formulaire rapide comme alternative
    
    if request.method == 'POST':
        # Traitement de l'import Excel
        if 'import_excel' in request.POST:
            if 'excel_file' in request.FILES:
                excel_file = request.FILES['excel_file']
                try:
                    # Vérifier l'extension du fichier
                    if not excel_file.name.endswith(('.xlsx', '.xls')):
                        messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls)")
                        return redirect('liste_eleves')
                    
                    # Charger le fichier Excel
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    
                    # Colonnes attendues
                    required_columns = ['Nom', 'Prénom']
                    optional_columns = ['Classe', 'Date de naissance', 'Téléphone', 'Email', 'Adresse']
                    expected_columns = required_columns + optional_columns

                    import unicodedata
                    def normalize(s):
                        if not s:
                            return ''
                        return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii').strip().lower()

                    # Normaliser les en-têtes du fichier
                    headers = [cell.value for cell in ws[1]]
                    norm_headers = [normalize(h) for h in headers]
                    norm_expected = [normalize(c) for c in expected_columns]

                    # Vérifier les colonnes obligatoires manquantes
                    missing_required = [col for col in required_columns if normalize(col) not in norm_headers]
                    if missing_required:
                        messages.error(request, f"Colonnes obligatoires manquantes dans le fichier Excel : {', '.join(missing_required)}.\nColonnes trouvées : {headers}")
                        return redirect('liste_eleves')

                    # Mapper les indices des colonnes (uniquement pour les colonnes présentes dans le fichier)
                    column_indices = {}
                    for col in expected_columns:
                        try:
                            column_indices[col] = norm_headers.index(normalize(col))
                        except ValueError:
                            # La colonne n'est pas dans le fichier, on l'ignore
                            pass
                    
                    # Compteurs pour le suivi
                    count_created = 0
                    count_skipped = 0
                    errors = []
                    
                    # Parcourir les lignes (en commençant par la deuxième ligne)
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        # Récupérer les valeurs (vérifier que les colonnes obligatoires sont présentes)
                        try:
                            nom = row[column_indices['Nom']].value
                            prenom = row[column_indices['Prénom']].value
                        except KeyError as e:
                            errors.append(f"Ligne {row_idx}: Colonne manquante - {str(e)}")
                            count_skipped += 1
                            continue
                        
                        # Vérifier que nom et prénom sont présents
                        if not nom or not prenom:
                            errors.append(f"Ligne {row_idx}: Nom ou prénom manquant")
                            count_skipped += 1
                            continue
                        
                        # Récupérer les autres valeurs (toutes optionnelles sauf Nom et Prénom)
                        classe_nom = row[column_indices.get('Classe')].value if 'Classe' in column_indices else None
                        date_naissance_val = row[column_indices.get('Date de naissance')].value if 'Date de naissance' in column_indices else None
                        telephone = row[column_indices.get('Téléphone')].value if 'Téléphone' in column_indices else ''
                        email = row[column_indices.get('Email')].value if 'Email' in column_indices else ''
                        adresse = row[column_indices.get('Adresse')].value if 'Adresse' in column_indices else ''
                        
                        # Initialiser la date de naissance par défaut
                        date_naissance = None
                        # Convertir la date de naissance si présente
                        if date_naissance_val:
                            if isinstance(date_naissance_val, datetime.datetime):
                                date_naissance = date_naissance_val.date()
                            elif isinstance(date_naissance_val, str):
                                try:
                                    # Essayer différents formats de date
                                    formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
                                    for fmt in formats:
                                        try:
                                            date_naissance = datetime.datetime.strptime(date_naissance_val, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                                    if not date_naissance:
                                        raise ValueError(f"Format de date non reconnu: {date_naissance_val}")
                                except ValueError as e:
                                    errors.append(f"Ligne {row_idx}: {str(e)}")
                                    count_skipped += 1
                                    continue
                        
                        # Vérifier si l'élève existe déjà dans la même composante (nom, prénom ET date de naissance)
                        if date_naissance is not None:
                            existe = Eleve.objects.filter(
                                nom__iexact=str(nom).strip(),
                                prenom__iexact=str(prenom).strip(),
                                date_naissance=date_naissance,
                                composante_id=composante_id
                            )
                        else:
                            existe = Eleve.objects.filter(
                                nom__iexact=str(nom).strip(),
                                prenom__iexact=str(prenom).strip(),
                                date_naissance__isnull=True,
                                composante_id=composante_id
                            )
                        if existe.exists():
                            print('[DEBUG] Doublon détecté:', list(existe.values('id', 'nom', 'prenom', 'date_naissance')))
                            errors.append(f"Ligne {row_idx}: L'élève {nom} {prenom} ({date_naissance if date_naissance else 'date inconnue'}) existe déjà")
                            count_skipped += 1
                            continue
                        
                        # Trouver la classe et le créneau correspondants
                        classe_obj = None
                        if classe_nom:
                            try:
                                # Recherche exacte d'abord
                                classe_obj = Classe.objects.get(nom=classe_nom, composante_id=composante_id)
                            except Classe.DoesNotExist:
                                try:
                                    # Recherche insensible à la casse
                                    classe_obj = Classe.objects.get(nom__iexact=classe_nom, composante_id=composante_id)
                                except Classe.DoesNotExist:
                                    try:
                                        # Recherche avec remplacement espace par underscore
                                        classe_nom_underscore = classe_nom.replace(' ', '_')
                                        classe_obj = Classe.objects.get(nom__iexact=classe_nom_underscore, composante_id=composante_id)
                                    except Classe.DoesNotExist:
                                        try:
                                            # Recherche sans filtre composante (au cas où)
                                            classe_obj = Classe.objects.get(nom__iexact=classe_nom)
                                        except Classe.DoesNotExist:
                                            try:
                                                # Recherche avec underscore sans composante
                                                classe_obj = Classe.objects.get(nom__iexact=classe_nom_underscore)
                                            except Classe.DoesNotExist:
                                                # Debug: afficher les classes disponibles (toutes les classes, pas seulement de la composante)
                                                classes_disponibles = list(Classe.objects.all().values_list('nom', flat=True))
                                                classes_composante = list(Classe.objects.filter(composante_id=composante_id).values_list('nom', flat=True))
                                                errors.append(f"Ligne {row_idx}: Classe '{classe_nom}' non trouvée. Classes de cette composante: {classes_composante}. Toutes classes: {classes_disponibles}")
                                                print(f"[DEBUG] Classe '{classe_nom}' non trouvée pour l'élève {nom} {prenom}")
                                                print(f"[DEBUG] Composante ID: {composante_id}")
                                                print(f"[DEBUG] Classes de cette composante: {classes_composante}")
                                                print(f"[DEBUG] Toutes les classes: {classes_disponibles}")
                        
                        creneau = None  # Le créneau n'est plus utilisé
                        
                        # Créer un utilisateur pour cet élève
                        username = generate_username('eleve', f"{nom}{prenom}")
                        password = generate_password(nom)  # Utiliser nom et prénom pour le mot de passe
                        
                        # Vérifier si le nom d'utilisateur existe déjà
                        while User.objects.filter(username=username).exists():
                            username = f"{username}{random.randint(1, 999)}"
                        
                        # Créer l'utilisateur
                        user = User.objects.create_user(username=username, password=password)
                        
                        # Créer l'élève
                        eleve = Eleve.objects.create(
                            nom=nom,
                            prenom=prenom,
                            date_naissance=date_naissance,
                            telephone=telephone,
                            email=email,
                            adresse=adresse,
                            user=user,
                            composante_id=composante_id,
                            mot_de_passe_en_clair=password
                        )
                        
                        # Assigner la classe si trouvée (relation ManyToMany)
                        if classe_obj:
                            eleve.classes.add(classe_obj)
                            print(f"[DEBUG] Élève {nom} {prenom} assigné à la classe {classe_obj.nom}")
                        else:
                            print(f"[DEBUG] Aucune classe assignée pour {nom} {prenom}")
                        count_created += 1
                    
                    # Afficher un résumé
                    # Affichage du message : si erreur, afficher uniquement error (barre rouge)
                    if errors:
                        messages.error(request, "Erreurs lors de l'import : " + "; ".join(errors[:5]) + ("... et d'autres" if len(errors) > 5 else ""))
                    else:
                        if count_skipped > 0:
                            messages.warning(request, f"{count_skipped} élève(s) ignoré(s) lors de l'import.")
                        elif count_created > 0:
                            messages.success(request, f"{count_created} élève(s) importé(s) avec succès !")
                    return redirect('liste_eleves')
                    
                except Exception as e:
                    import traceback
                    print('[IMPORT EXCEL][ERREUR]', str(e))
                    traceback.print_exc()
                    messages.error(request, f"Erreur lors de l'import du fichier Excel : {str(e)}")
                    return redirect('liste_eleves')
            else:
                messages.error(request, "Aucun fichier Excel n'a été fourni")
                return redirect('liste_eleves')
        
        # Vérifier si l'élève avec le même nom, prénom ET date de naissance existe déjà
        nom = request.POST.get('nom', '').strip()
        prenom = request.POST.get('prenom', '').strip()
        date_naissance = request.POST.get('date_naissance', '').strip()
        date_obj = None
        if date_naissance:
            try:
                date_obj = datetime.strptime(date_naissance, "%Y-%m-%d").date()
            except Exception:
                date_obj = None
        if date_obj:
            existe = Eleve.objects.filter(
                nom__iexact=str(nom).strip(),
                prenom__iexact=str(prenom).strip(),
                date_naissance=date_obj
            )
        else:
            existe = Eleve.objects.filter(
                nom__iexact=str(nom).strip(),
                prenom__iexact=str(prenom).strip(),
                date_naissance__isnull=True
            )
        if existe.exists():
            print('[DEBUG] Doublon détecté:', list(existe.values('id', 'nom', 'prenom', 'date_naissance')))
            messages.warning(request, f'Un élève avec le nom "{nom} {prenom}" et cette date de naissance existe déjà!')
            return redirect('liste_eleves')
            
        if 'ajout_complet' in request.POST:
            form = EleveForm(request.POST, request=request)
            if form.is_valid():
                nom = form.cleaned_data['nom']
                prenom = form.cleaned_data['prenom']
                classe = form.cleaned_data.get('classe')
                if classe and classe.eleves.count() >= classe.capacite:
                    messages.error(request, f"La classe {classe} est déjà pleine ({classe.capacite} élèves) !")
                    return redirect('liste_eleves')
                # Créer un utilisateur pour cet élève
                username = generate_username('eleve', f"{nom}{prenom}")
                password = generate_password(f"{nom} {prenom}")
                
                # Vérifier si le nom d'utilisateur existe déjà
                while User.objects.filter(username=username).exists():
                    username = f"{username}{random.randint(1, 999)}"
                
                # Créer l'utilisateur
                user = User.objects.create_user(username=username, password=password)
                
                # Créer l'élève et l'associer à l'utilisateur
                eleve = form.save(commit=False)
                eleve.user = user
                eleve.composante_id = composante_id
                eleve.mot_de_passe_en_clair = password
                eleve.save()
                
                # Enregistrer les relations many-to-many
                form.save_m2m()
                
                messages.success(request, f"L'élève {eleve.prenom} {eleve.nom} a été ajouté avec succès ! Identifiants : {username} / {password}")
                return redirect('liste_eleves')
        
        elif 'ajout_rapide' in request.POST:
            form_rapide = EleveRapideForm(request.POST)
            if form_rapide.is_valid():
                nom = form_rapide.cleaned_data['nom']
                prenom = form_rapide.cleaned_data['prenom']
                classe = form_rapide.cleaned_data.get('classe')
                
                if classe and classe.eleves.count() >= classe.capacite:
                    messages.error(request, f"La classe {classe} est déjà pleine ({classe.capacite} élèves) !")
                    return redirect('liste_eleves')
                
                # Créer un utilisateur pour cet élève
                username = generate_username('eleve', f"{nom}{prenom}")
                password = generate_password(f"{nom} {prenom}")
                
                # Vérifier si le nom d'utilisateur existe déjà
                while User.objects.filter(username=username).exists():
                    username = f"{username}{random.randint(1, 999)}"
                
                # Créer l'utilisateur
                user = User.objects.create_user(username=username, password=password)
                
                # Créer l'élève
                eleve = Eleve(
                    nom=nom,
                    prenom=prenom,
                    classe=classe,
                    composante_id=composante_id,
                    user=user,
                    mot_de_passe_en_clair=password,
                    montant_total=200.00  # Montant par défaut
                )
                eleve.save()
                
                messages.success(request, f"L'élève {prenom} {nom} a été ajouté rapidement avec succès ! Identifiants : {username} / {password}")
                return redirect('liste_eleves')
    
    context = {
        'eleves': eleves,
        'form': form,
        'form_rapide': form_rapide,
        'classes': Classe.objects.all(),
        'creneaux': Creneau.objects.all(),
        'selected_classe': int(classe_id) if classe_id else None,
        'selected_creneau': int(creneau_id) if creneau_id else None,
    }
    
    return render(request, 'ecole_app/eleves/liste.html', context)

@login_required
def detail_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    form = EleveForm(instance=eleve)
    
    # Récupérer les paiements de l'élève
    paiements = Paiement.objects.filter(eleve=eleve).order_by('-date')
    
    # Les montants payés et restants sont calculés via les propriétés du modèle Eleve
    # Pas besoin de les calculer ou de les assigner ici
    
    # Récupérer les informations d'identification si l'utilisateur existe
    user_info = None
    if eleve.user:
        user_info = {
            'username': eleve.user.username,
            'mot_de_passe_en_clair': eleve.mot_de_passe_en_clair,
            # On ne stocke pas le mot de passe en clair, nous le générerons lors de l'envoi par email
        }
        
        # Vérifier si un mot de passe temporaire est disponible dans la session
        temp_password = None
        if 'temp_password' in request.session and 'eleve_id' in request.session and str(request.session['eleve_id']) == str(eleve_id):
            temp_password = request.session['temp_password']
            # On supprime le mot de passe de la session après l'avoir récupéré
            del request.session['temp_password']
            del request.session['eleve_id']
            request.session.modified = True
            
            # Ajouter le mot de passe temporaire aux informations utilisateur
            user_info['temp_password'] = temp_password
    
    context = {
        'eleve': eleve,
        'form': form,
        'paiements': paiements,
        'user_info': user_info,
    }
    
    return render(request, 'ecole_app/eleves/detail.html', context)


@login_required
def coran_eleve(request, eleve_id):
    import os
    import json
    from django.conf import settings
    eleve = get_object_or_404(Eleve, id=eleve_id)
    # Chemin vers le fichier JSON du Coran
    coran_path = os.path.join(settings.BASE_DIR, 'ecole_app', 'static', 'coran', 'quran_ar.json')
    with open(coran_path, encoding='utf-8') as f:
        coran_data = json.load(f)
    # Lire la liste complète des sourates depuis le fichier JSON
    sourates_path = os.path.join(settings.BASE_DIR, 'ecole_app', 'static', 'coran', 'sourates.json')
    with open(sourates_path, encoding='utf-8') as f:
        sourates = json.load(f)
    # Récupérer la sourate sélectionnée
    sourate_num = int(request.GET.get('sourate', 1))
    # Filtrer les versets de la sourate
    versets = [v for v in coran_data['quran'] if v['sura'] == sourate_num]
    context = {
        'eleve': eleve,
        'sourates': sourates,
        'sourate_active': sourate_num,
        'versets': versets,
    }
    return render(request, 'ecole_app/eleves/coran.html', context)

@login_required
@require_POST
def supprimer_eleve(request, eleve_id):
    from django.core.exceptions import ObjectDoesNotExist
    try:
        eleve = Eleve.objects.get(id=eleve_id)
        nom = str(eleve)
        eleve.delete()
        messages.success(request, f"L'élève {nom} a été supprimé avec succès !")
    except Eleve.DoesNotExist:
        messages.error(request, "L'élève demandé n'existe plus ou a déjà été supprimé.")
    return redirect('liste_eleves')

# Gestion des professeurs
@login_required
def liste_professeurs(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la liste des professeurs.")
        return redirect('selection_composante')
        
    # Filtrer les professeurs par composante (relation ManyToMany)
    from .models import Composante
    composante = get_object_or_404(Composante, id=composante_id)
    professeurs = Professeur.objects.filter(composantes=composante).order_by('nom')
    form = ProfesseurForm(request=request)
    
    if request.method == 'POST':
        form = ProfesseurForm(request.POST, request=request)
        
        if form.is_valid():
            # Récupérer le nom du formulaire
            nom = request.POST.get('nom_complet', '').strip()
            
            if not nom:
                messages.error(request, "Le nom du professeur ne peut pas être vide.")
            else:
                # Vérifier si un professeur avec ce nom existe déjà
                professeur_existant = Professeur.objects.filter(nom__iexact=nom).first()
                
                if professeur_existant:
                    # Le professeur existe déjà, l'ajouter à cette composante s'il n'y est pas déjà
                    if not professeur_existant.composantes.filter(id=composante_id).exists():
                        professeur_existant.composantes.add(composante)
                        messages.success(request, f"Le professeur '{nom}' a été ajouté à cette composante.")
                    else:
                        messages.warning(request, f"Le professeur '{nom}' est déjà dans cette composante.")
                else:
                    # Créer un nouveau professeur
                    # Créer un utilisateur pour ce professeur
                    username = generate_username(nom, '')
                    password = generate_password(nom)
                    
                    # Vérifier si le nom d'utilisateur existe déjà
                    while User.objects.filter(username=username).exists():
                        username = f"{username}{random.randint(1, 999)}"
                    
                    # Créer l'utilisateur
                    user = User.objects.create_user(username=username, password=password)
                    
                    # Créer le professeur directement sans passer par le formulaire
                    # pour éviter les problèmes de many-to-many
                    indemnisation_value = form.cleaned_data.get('indemnisation', 0.0)
                    if indemnisation_value is None:
                        indemnisation_value = 0.0
                    
                    professeur = Professeur(
                        user=user,
                        nom=form.cleaned_data.get('nom_complet', '').strip(),
                        email=form.cleaned_data.get('email', ''),
                        telephone=form.cleaned_data.get('telephone', ''),
                        indemnisation=indemnisation_value,
                        mot_de_passe_en_clair=password  # Stocker le mot de passe en clair
                    )
                    professeur.save()  # Sauvegarder d'abord pour obtenir un ID
                    
                    # Ajouter toutes les composantes au professeur
                    toutes_composantes = Composante.objects.all()
                    if toutes_composantes.exists():
                        for comp in toutes_composantes:
                            professeur.composantes.add(comp)
                    
                    # Afficher les informations de connexion
                    messages.success(request, f'Le professeur a été créé avec succès ! Identifiant: {username} | Mot de passe: {password}')
                    
                return redirect('liste_professeurs')
    
    context = {
        'professeurs': professeurs,
        'form': form,
        'composante': composante,
    }
    
    return render(request, 'ecole_app/professeurs/liste.html', context)

@login_required
def detail_professeur(request, professeur_id):
    professeur = get_object_or_404(Professeur, id=professeur_id)
    classes = Classe.objects.filter(professeur=professeur)
    
    if request.method == 'POST':
        form = ProfesseurForm(request.POST, instance=professeur, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Le professeur a été mis à jour avec succès !')
            return redirect('liste_professeurs')
    else:
        form = ProfesseurForm(instance=professeur, initial={'nom_complet': professeur.nom}, request=request)
    
    # Récupérer les informations d'identification si l'utilisateur existe
    user_info = None
    if professeur.user:
        user_info = {
            'username': professeur.user.username,
            'mot_de_passe_en_clair': professeur.mot_de_passe_en_clair,
        }
        
        # Vérifier si un mot de passe temporaire est disponible dans la session
        temp_password = None
        if 'temp_password' in request.session and 'professeur_id' in request.session and str(request.session['professeur_id']) == str(professeur_id):
            temp_password = request.session['temp_password']
            # On supprime le mot de passe de la session après l'avoir récupéré
            del request.session['temp_password']
            del request.session['professeur_id']
            request.session.modified = True
            
            # Ajouter le mot de passe temporaire aux informations utilisateur
            user_info['temp_password'] = temp_password
    
    context = {
        'professeur': professeur,
        'form': form,
        'classes': classes,
        'user_info': user_info,
    }
    
    return render(request, 'ecole_app/professeurs/detail.html', context)

@login_required
@require_POST
def supprimer_professeur(request, professeur_id):
    from django.core.exceptions import ObjectDoesNotExist
    try:
        professeur = Professeur.objects.get(id=professeur_id)
        nom = professeur.nom
        professeur.delete()
        messages.success(request, f'Le professeur {nom} a été supprimé avec succès !')
    except Professeur.DoesNotExist:
        messages.error(request, "Le professeur demandé n'existe plus ou a déjà été supprimé.")
    return redirect('liste_professeurs')

# Gestion des classes
@login_required
def liste_classes(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la liste des classes.")
        return redirect('selection_composante')
        
    classes = Classe.objects.filter(composante_id=composante_id).order_by('nom')
    form = ClasseForm(request=request)
    
    if request.method == 'POST':
        form = ClasseForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'La classe a été ajoutée avec succès !')
            return redirect('liste_classes')
    
    context = {
        'classes': classes,
        'form': form,
    }
    
    return render(request, 'ecole_app/classes/liste.html', context)

@login_required
def detail_classe(request, classe_id):
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Récupérer les élèves des deux relations (ForeignKey et ManyToMany)
    eleves_fk = classe.eleves.filter(archive=False)
    eleves_m2m = classe.eleves_multi.filter(archive=False)
    
    # Fusionner les deux querysets sans doublons
    eleves = (eleves_fk | eleves_m2m).distinct().order_by('nom', 'prenom')
    
    if request.method == 'POST':
        form = ClasseForm(request.POST, instance=classe, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Les informations ont été mises à jour avec succès !')
            return redirect('detail_classe', classe_id=classe.id)
    else:
        form = ClasseForm(instance=classe, request=request)
    
    context = {
        'classe': classe,
        'form': form,
        'eleves': eleves,
        'taux_occupation': classe.taux_occupation(),
    }
    
    return render(request, 'ecole_app/classes/detail.html', context)

@login_required
@require_POST
def supprimer_classe(request, classe_id):
    classe = get_object_or_404(Classe, id=classe_id)
    nom = classe.nom
    classe.delete()
    messages.success(request, f'La classe {nom} a été supprimée avec succès !')
    return redirect('liste_classes')

# Gestion des créneaux
@login_required
def detail_creneau(request, creneau_id):
    creneau = get_object_or_404(Creneau, id=creneau_id)
    classes = Classe.objects.filter(creneau=creneau)
    eleves = Eleve.objects.filter(creneaux=creneau)
    context = {
        'creneau': creneau,
        'classes': classes,
        'eleves': eleves,
    }
    return render(request, 'ecole_app/creneaux/detail.html', context)

@login_required
def liste_creneaux(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la liste des créneaux.")
        return redirect('selection_composante')
        
    creneaux = Creneau.objects.filter(composante_id=composante_id).order_by('nom')
    form = CreneauForm(request=request)
    
    if request.method == 'POST':
        form = CreneauForm(request.POST, request=request)
        if form.is_valid():
            nom = form.cleaned_data.get('nom', '').strip()
            if not nom:
                messages.error(request, "Le nom du créneau est obligatoire.")
            elif Creneau.objects.filter(nom__iexact=nom, composante_id=composante_id).exists():
                messages.error(request, f"Un créneau nommé '{nom}' existe déjà dans cette composante.")
            else:
                form.save()
                messages.success(request, 'Le créneau a été ajouté avec succès !')
                return redirect('liste_creneaux')
    
    context = {
        'creneaux': creneaux,
        'form': form,
    }
    
    return render(request, 'ecole_app/creneaux/liste.html', context)

@login_required
@require_POST
def supprimer_creneau(request, creneau_id):
    from django.core.exceptions import ObjectDoesNotExist
    try:
        creneau = Creneau.objects.get(id=creneau_id)
        nom = creneau.nom
        creneau.delete()
        messages.success(request, f'Le créneau {nom} a été supprimé avec succès !')
    except Creneau.DoesNotExist:
        messages.error(request, "Le créneau demandé n'existe plus ou a déjà été supprimé.")
    return redirect('liste_creneaux')

# Gestion des paiements
@login_required
def liste_paiements(request):
    paiements = Paiement.objects.all().order_by('-date')
    form = PaiementForm()
    
    if request.method == 'POST':
        form = PaiementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Le paiement a été enregistré avec succès !')
            return redirect('liste_paiements')
    
    # Statistiques financières
    total_paiements = paiements.aggregate(Sum('montant'))['montant__sum'] or 0
    nombre_eleves = Eleve.objects.count()
    moyenne_paiement = total_paiements / nombre_eleves if nombre_eleves > 0 else 0

    # Statistiques financières globales (tous les élèves et paiements)
    from .models import Eleve, Paiement
    total_attendu = Eleve.objects.aggregate(total=Sum('montant_total'))['total'] or 0
    total_collecte = Paiement.objects.aggregate(total=Sum('montant'))['total'] or 0
    pourcentage_collecte = round((total_collecte / total_attendu) * 100) if total_attendu > 0 else 0

    context = {
        'paiements': paiements,
        'form': form,
        'total_paiements': total_paiements,
        'nombre_eleves': nombre_eleves,
        'moyenne_paiement': moyenne_paiement,
        'total_attendu': total_attendu,
        'total_collecte': total_collecte,
        'pourcentage_collecte': pourcentage_collecte,
    }
    
    return render(request, 'ecole_app/paiements/liste.html', context)

@login_required
def supprimer_paiement(request, paiement_id):
    try:
        paiement = Paiement.objects.get(id=paiement_id)
        eleve = paiement.eleve
        paiement.delete()
        messages.success(request, 'Le paiement a été supprimé avec succès !')
        # Rediriger vers la liste des paiements si l'élève n'existe plus
        if eleve:
            return redirect('detail_eleve', eleve_id=eleve.id)
        else:
            return redirect('liste_paiements')
    except Paiement.DoesNotExist:
        messages.warning(request, 'Ce paiement n’existe plus ou a déjà été supprimé.')
        return redirect('liste_paiements')

@login_required
def export_paiements(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="paiements-{datetime.datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Élève', 'Montant', 'Date de paiement', 'Méthode', 'Commentaire'])
    
    for paiement in Paiement.objects.all().order_by('-date'):
        writer.writerow([
            paiement.id, 
            f"{paiement.eleve.nom} {paiement.eleve.prenom}", 
            paiement.montant, 
            paiement.date.strftime('%d/%m/%Y'), 
            paiement.methode, 
            paiement.commentaire
        ])
    
    return response

# Import/Export de données
@login_required
def import_export(request):
    import_form = ImportDataForm()
    export_form = ExportDataForm()
    
    context = {
        'import_form': import_form,
        'export_form': export_form,
    }
    
    return render(request, 'ecole_app/import_export.html', context)

@login_required
@require_POST
def import_data(request):
    form = ImportDataForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            fichier = request.FILES['fichier_import']
            type_fichier = form.cleaned_data['type_fichier']
            count = 0
            if type_fichier == 'csv':
                decoded_file = fichier.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                for row in reader:
                    Eleve.objects.create(
                        nom=row.get('nom', ''),
                        prenom=row.get('prenom', ''),
                        classe_id=row.get('classe_id') or None,
                        creneau_id=row.get('creneau_id') or None,
                        date_naissance=row.get('date_naissance') or None,
                        telephone=row.get('telephone', ''),
                        email=row.get('email', ''),
                        adresse=row.get('adresse', '')
                    )
                    count += 1
            elif type_fichier == 'excel':
                wb = openpyxl.load_workbook(fichier)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    Eleve.objects.create(
                        nom=data.get('nom', ''),
                        prenom=data.get('prenom', ''),
                        classe_id=data.get('classe_id') or None,
                        creneau_id=data.get('creneau_id') or None,
                        date_naissance=data.get('date_naissance') or None,
                        telephone=data.get('telephone', ''),
                        email=data.get('email', ''),
                        adresse=data.get('adresse', '')
                    )
                    count += 1
            messages.success(request, f'{count} élèves ont été importés avec succès !')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import : {str(e)}')
    return redirect('import_export')

@login_required
def export_data(request):
    type_export = request.GET.get('type_export', 'eleves')
    format_export = request.GET.get('format_export', 'excel')
    data = {}

    if type_export in ['eleves', 'tout']:
        data['eleves'] = list(Eleve.objects.values())
    if type_export in ['professeurs', 'tout']:
        data['professeurs'] = list(Professeur.objects.values())
    if type_export in ['classes', 'tout']:
        data['classes'] = list(Classe.objects.values())
    if type_export in ['creneaux', 'tout']:
        data['creneaux'] = list(Creneau.objects.values())
    if type_export in ['paiements', 'tout']:
        data['paiements'] = list(Paiement.objects.values())

    if format_export == 'json':
        response = HttpResponse(json.dumps(data, default=str, indent=2), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="donnees-ecole-{datetime.datetime.now().strftime("%Y%m%d")}.json"'
        return response
    elif format_export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="donnees-ecole-{datetime.datetime.now().strftime("%Y%m%d")}.csv"'
        writer = csv.writer(response)
        if type_export == 'eleves' or type_export == 'tout':
            writer.writerow(['ID', 'Nom', 'Prénom', 'Classe', 'Créneau', 'Date de naissance', 'Téléphone', 'Email', 'Adresse'])
            for eleve in Eleve.objects.all():
                writer.writerow([
                    eleve.id, eleve.nom, eleve.prenom, 
                    eleve.classe.nom if eleve.classe else '', 
                    eleve.creneau.nom if eleve.creneau else '',
                    eleve.date_naissance, eleve.telephone, eleve.email, eleve.adresse
                ])
        return response
    elif format_export == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Élèves'
        ws.append(['ID', 'Nom', 'Prénom', 'Classe', 'Créneau', 'Date de naissance', 'Téléphone', 'Email', 'Adresse'])
        for eleve in Eleve.objects.all():
            ws.append([
                eleve.id, eleve.nom, eleve.prenom,
                eleve.classe.nom if eleve.classe else '',
                eleve.creneau.nom if eleve.creneau else '',
                eleve.date_naissance, eleve.telephone, eleve.email, eleve.adresse
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="eleves-{datetime.datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    else:
        messages.error(request, "Format d'export non supporté.")
        return redirect('import_export')

@login_required
def modifier_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    print(f"[DEBUG] modifier_eleve appelé pour l'élève {eleve_id}")
    
    if request.method == 'POST':
        print(f"[DEBUG] Méthode POST détectée dans modifier_eleve")
        form = EleveForm(request.POST, instance=eleve, request=request)
        if form.is_valid():
            print(f"[DEBUG] Formulaire valide, sauvegarde en cours...")
            form.save()
            messages.success(request, 'Les informations de l\'élève ont été mises à jour avec succès !')
            return redirect('liste_eleves')
        else:
            print(f"[DEBUG] Formulaire NON valide dans modifier_eleve")
            print(f"[DEBUG] Erreurs: {form.errors.as_json()}")
    else:
        print(f"[DEBUG] Méthode GET détectée dans modifier_eleve")
        form = EleveForm(instance=eleve, request=request)
    
    context = {
        'form': form,
        'eleve': eleve,
        'mode_edition': True
    }
    
    return render(request, 'ecole_app/eleves/modifier.html', context)

# Ajout rapide d'élèves (vue simplifiée)
@login_required
def ajout_rapide_eleve(request):
    if request.method == 'POST':
        form = EleveRapideForm(request.POST, request=request)
        if form.is_valid():
            eleve = form.save()
            messages.success(request, f'L\'élève {eleve} a été ajouté avec succès !')
            # Redirection vers le même formulaire pour ajouter un autre élève
            return redirect('ajout_rapide_eleve')
    else:
        form = EleveRapideForm(request=request)
    
    context = {
        'form': form,
        'classes': Classe.objects.all(),
        'creneaux': Creneau.objects.all(),
    }
    
    return render(request, 'ecole_app/eleves/ajout_rapide.html', context)

# Vue pour la liste des paiements
@login_required
def liste_paiements(request):
    # Vérifier si une composante est sélectionnée
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, 'Veuillez sélectionner une composante pour continuer.')
        return redirect('selection_composante')
        
    # Filtrer les paiements par composante active
    paiements = Paiement.objects.filter(composante_id=composante_id).order_by('-date')
    form = PaiementForm(request=request)
    
    if request.method == 'POST':
        form = PaiementForm(request.POST, request=request)
        if form.is_valid():
            # Récupérer les données du formulaire
            eleve = form.cleaned_data['eleve']
            montant = form.cleaned_data['montant']
            date = form.cleaned_data['date']
            methode = form.cleaned_data['methode']
            commentaire = form.cleaned_data['commentaire']
            annee_scolaire = form.cleaned_data['annee_scolaire']
            
            # Vérifier si un paiement existe déjà pour cet élève
            paiement_existant = Paiement.objects.filter(
                eleve=eleve,
                composante_id=composante_id
            ).first()
            
            from .models import PaiementHistorique
            if paiement_existant:
                # Mettre à jour la date et la méthode du paiement existant
                paiement_existant.date = date  # Mettre à jour avec la date la plus récente
                paiement_existant.methode = methode
                
                # Ajouter le nouveau commentaire au commentaire existant s'il y en a un
                if commentaire:
                    if paiement_existant.commentaire:
                        paiement_existant.commentaire += f" | {date.strftime('%d/%m/%Y')}: {commentaire}"
                    else:
                        paiement_existant.commentaire = f"{date.strftime('%d/%m/%Y')}: {commentaire}"
                paiement_existant.save()
                
                # Ajouter une entrée historique
                PaiementHistorique.objects.create(
                    paiement=paiement_existant,
                    montant=montant,
                    date=date,
                    methode=methode,
                    commentaire=commentaire or ''
                )
                
                # Recalculer le montant total du paiement à partir des historiques
                recalculer_montant_paiement(paiement_existant)
                
                messages.success(request, f'Le paiement pour {eleve} a été mis à jour avec succès!')
            else:
                # Créer un nouveau paiement
                nouveau_paiement = form.save(commit=False)
                nouveau_paiement.composante_id = composante_id
                nouveau_paiement.montant = 0  # Initialiser à 0, sera recalculé
                nouveau_paiement.save()
                
                # Ajouter une entrée historique
                PaiementHistorique.objects.create(
                    paiement=nouveau_paiement,
                    montant=montant,
                    date=date,
                    methode=methode,
                    commentaire=commentaire or ''
                )
                
                # Recalculer le montant total du paiement à partir des historiques
                recalculer_montant_paiement(nouveau_paiement)
                
                messages.success(request, 'Le paiement a été enregistré avec succès!')
    
    # Calcul du montant total collecté
    total_collecte = paiements.aggregate(Sum('montant'))['montant__sum'] or 0
    
    # Récupérer le montant total attendu (somme des montants personnalisés de chaque élève)
    from .models import Eleve
    total_attendu = Eleve.objects.filter(composante_id=composante_id).aggregate(total=Sum('montant_total'))['total'] or 0
    
    # Calcul du pourcentage collecté
    pourcentage_collecte = round((total_collecte / total_attendu * 100), 2) if total_attendu > 0 else 0
    
    context = {
        'paiements': paiements,
        'form': form,
        'total_collecte': total_collecte,
        'total_attendu': total_attendu,
        'pourcentage_collecte': pourcentage_collecte,
        'now': timezone.now()  # Pour afficher la date de dernière mise à jour
    }
    
    return render(request, 'ecole_app/paiements/liste.html', context)

# Vue pour le détail d'un paiement
@login_required
def detail_paiement(request, paiement_id):
    paiement = get_object_or_404(Paiement, id=paiement_id)
    autres_paiements = Paiement.objects.filter(eleve=paiement.eleve).exclude(id=paiement_id).order_by('-date')

    if request.method == 'POST':
        form = PaiementForm(request.POST, instance=paiement, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Le paiement a été modifié avec succès !')
            return redirect('detail_paiement', paiement_id=paiement.id)
        else:
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire.')
    else:
        form = PaiementForm(instance=paiement, request=request)

    # Calcul du total payé par l'élève
    total_paiements = Paiement.objects.filter(eleve=paiement.eleve).aggregate(Sum('montant'))['montant__sum'] or 0

    historiques = paiement.historiques.order_by('date', 'date_creation')
    context = {
        'paiement': paiement,
        'form': form,
        'autres_paiements': autres_paiements,
        'total_paiements': total_paiements,
        'historiques': historiques,
    }
    return render(request, 'ecole_app/paiements/detail.html', context)

# Vue pour exporter les paiements
@login_required
def export_paiements(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="paiements-{datetime.datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"
    
    # Définir les en-têtes
    headers = ['ID', 'Élève', 'Date', 'Montant', 'Méthode de paiement', 'Commentaire']
    ws.append(headers)
    
    # Appliquer le style aux en-têtes
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    # Ajouter les données
    paiements = Paiement.objects.all().order_by('-date')
    for paiement in paiements:
        ws.append([
            paiement.id,
            f"{paiement.eleve.nom} {paiement.eleve.prenom}",
            paiement.date,
            paiement.montant,
            paiement.get_methode_display(),
            paiement.commentaire or ""
        ])
    
    wb.save(response)
    return response

# Vue pour supprimer un paiement
@login_required
def supprimer_paiement(request, paiement_id):
    paiement = get_object_or_404(Paiement, id=paiement_id)
    if request.method == 'POST':
        paiement.delete()
        messages.success(request, 'Le paiement a été supprimé avec succès!')
        return redirect('liste_paiements')
    return redirect('liste_paiements')

# Vue pour régénérer le mot de passe d'un élève
@login_required
def regenerer_password_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    
    # Vérifier si l'élève a un compte utilisateur associé
    if not eleve.user:
        messages.error(request, "Cet élève n'a pas de compte utilisateur associé.")
        return redirect('detail_eleve', eleve_id=eleve.id)
    
    # Générer un nouveau mot de passe
    password = generate_password(eleve.nom, eleve.prenom)
    eleve.user.set_password(password)
    eleve.user.save()
    
    # Stocker le mot de passe temporairement dans la session pour l'affichage
    request.session['temp_password'] = password
    request.session['temp_username'] = eleve.user.username
    request.session['eleve_id'] = eleve.id
    
    messages.success(request, f"Nouveau mot de passe généré avec succès !")
    return redirect('detail_eleve', eleve_id=eleve.id)


# Vue pour régénérer le mot de passe d'un professeur
@login_required
def regenerer_password_professeur(request, professeur_id):
    professeur = get_object_or_404(Professeur, id=professeur_id)
    
    # Vérifier si le professeur a un compte utilisateur associé
    if not professeur.user:
        messages.error(request, "Ce professeur n'a pas de compte utilisateur associé.")
        return redirect('detail_professeur', professeur_id=professeur.id)
    
    # Générer un nouveau mot de passe
    password = generate_password(professeur.nom)
    professeur.user.set_password(password)
    professeur.user.save()
    
    # Stocker le mot de passe temporairement dans la session pour l'affichage
    request.session['temp_password'] = password
    request.session['temp_username'] = professeur.user.username
    request.session['professeur_id'] = professeur.id
    
    messages.success(request, f"Nouveau mot de passe généré avec succès !")
    return redirect('detail_professeur', professeur_id=professeur.id)


# Vue pour envoyer les identifiants d'un élève par email
@login_required
def envoyer_identifiants_eleve(request, eleve_id):
    from django.conf import settings
    eleve = get_object_or_404(Eleve, id=eleve_id)
    
    # Vérifier si l'élève a un compte utilisateur associé
    if not eleve.user:
        messages.error(request, "Cet élève n'a pas de compte utilisateur associé.")
        return redirect('detail_eleve', eleve_id=eleve.id)
    
    # Vérifier si l'email est fourni
    email = request.POST.get('email', '')
    if not email:
        messages.error(request, "Veuillez fournir une adresse email valide.")
        return redirect('detail_eleve', eleve_id=eleve.id)
    
    # Générer un nouveau mot de passe au format nom.prenom1
    nom = eleve.nom.lower() if eleve.nom else ''
    prenom = eleve.prenom.lower() if eleve.prenom else ''
    
    # Nettoyer les caractères spéciaux et espaces
    import re
    nom = re.sub(r'[^a-z0-9]', '', nom)
    prenom = re.sub(r'[^a-z0-9]', '', prenom)
    
    # Créer le mot de passe au format nom.prenom1
    password = f"{nom}.{prenom}1" if nom and prenom else generer_mot_de_passe()
    
    # Mettre à jour le mot de passe de l'utilisateur
    eleve.user.set_password(password)
    eleve.user.save()
    
    # Mettre à jour le mot de passe en clair pour l'affichage
    eleve.mot_de_passe_en_clair = password
    eleve.save()
    
    # Envoyer l'email avec les identifiants
    subject = "Vos identifiants de connexion - Al Markaz"
    
    # Message HTML avec des couleurs et un design élaboré
    html_message = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 600px;
                margin: 0 auto;
            }}
            .container {{
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                overflow: hidden;
                box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #1e5631, #3a7f4e);
                color: white;
                padding: 20px;
                text-align: center;
                font-size: 24px;
            }}
            .arabic {{
                font-family: 'Traditional Arabic', serif;
                font-size: 28px;
                text-align: center;
                margin: 20px 0;
                color: #1e5631;
            }}
            .content {{
                padding: 20px;
                background-color: #f9f9f9;
            }}
            .greeting {{
                font-size: 18px;
                margin-bottom: 20px;
            }}
            .credentials {{
                background-color: white;
                border-left: 4px solid #3a7f4e;
                padding: 15px;
                margin: 20px 0;
                border-radius: 5px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            }}
            .credential-item {{
                margin: 10px 0;
            }}
            .credential-label {{
                font-weight: bold;
                color: #1e5631;
                display: inline-block;
                width: 100px;
            }}
            .credential-value {{
                font-family: monospace;
                background-color: #f0f0f0;
                padding: 3px 8px;
                border-radius: 3px;
                border: 1px solid #ddd;
            }}
            .advice {{
                background-color: #fff8e1;
                border-left: 4px solid #ffc107;
                padding: 15px;
                margin: 20px 0;
                border-radius: 5px;
            }}
            .advice-icon {{
                font-size: 18px;
                margin-right: 10px;
            }}
            .quran {{
                background-color: #e8f5e9;
                border-left: 4px solid #4caf50;
                padding: 15px;
                margin: 20px 0;
                border-radius: 5px;
                font-style: italic;
            }}
            .quran-reference {{
                text-align: right;
                font-weight: bold;
                color: #1e5631;
            }}
            .footer {{
                background-color: #f5f5f5;
                padding: 15px;
                text-align: center;
                border-top: 1px solid #e0e0e0;
            }}
            .footer-text {{
                color: #666;
            }}
            .signature {{
                margin-top: 20px;
                color: #1e5631;
                font-weight: bold;
            }}
            .mosque-icon {{
                width: 100%;
                text-align: center;
                margin: 10px 0;
                font-size: 24px;
            }}
            .divider {{
                border-top: 1px solid #e0e0e0;
                margin: 20px 0;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                Al Markaz
            </div>
            <div class="content">
                <div class="arabic">بِسْمِ اللَّهِ الرَّحْمَـٰنِ الرَّحِيمِ</div>
                
                <div class="greeting">
                    Assalamu alaykum wa rahmatullahi wa barakatuh <strong>{eleve.prenom} {eleve.nom}</strong>,
                </div>
                
                <p>Nous vous transmettons vos identifiants de connexion :</p>
                
                <div class="credentials">
                    <div class="credential-item">
                        <span class="credential-label">📱 Identifiant</span>: 
                        <span class="credential-value">{eleve.user.username}</span>
                    </div>
                    <div class="credential-item">
                        <span class="credential-label">🔐 Mot de passe</span>: 
                        <span class="credential-value">{password}</span>
                    </div>
                </div>
                
                <div class="advice">
                    <span class="advice-icon">🌙</span> <strong>Conseil</strong>: Gardez ces informations confidentielles et changez votre mot de passe régulièrement.
                </div>
                
                <div class="mosque-icon">🕋</div>
                
                <div class="quran">
                    <p>"Et rappelle-toi ton Seigneur dans ton âme avec humilité et crainte, et à voix basse, le matin et le soir, et ne sois pas du nombre des négligents."</p>
                    <div class="quran-reference">Coran 7:205</div>
                </div>
                
                <div class="divider"></div>
                
                <div class="signature">
                    Barakallahu fikum,<br>
                    L'équipe Al Markaz
                </div>
            </div>
            <div class="footer">
                <div class="footer-text">© {datetime.datetime.now().year} Al Markaz - Tous droits réservés</div>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Message texte simple pour les clients qui ne supportent pas le HTML
    message = f"""بِسْمِ اللَّهِ الرَّحْمَـٰنِ الرَّحِيمِ

Assalamu alaykum wa rahmatullahi wa barakatuh {eleve.prenom} {eleve.nom},

Nous vous transmettons vos identifiants de connexion :

📱 Identifiant : {eleve.user.username}
🔐 Mot de passe : {password}

🌙 Conseil : Gardez ces informations confidentielles et changez votre mot de passe régulièrement.

📖 "Et rappelle-toi ton Seigneur dans ton âme avec humilité et crainte, et à voix basse, le matin et le soir, et ne sois pas du nombre des négligents." - Coran 7:205

Barakallahu fikum,
L'équipe Al Markaz
    """
    
    sender_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [email]
    
    # Afficher les identifiants directement dans l'interface
    messages.info(request, f"Identifiant: {eleve.user.username}")
    messages.info(request, f"Mot de passe: {password}")
    
    try:
        # Utilisation de EmailMultiAlternatives pour envoyer à la fois du texte et du HTML
        from django.core.mail import EmailMultiAlternatives
        
        # Création de l'email avec le texte simple
        email_message = EmailMultiAlternatives(subject, message, sender_email, recipient_list)
        
        # Ajout du contenu HTML comme alternative
        email_message.attach_alternative(html_message, "text/html")
        
        # Envoi de l'email
        email_message.send()
        
        messages.success(request, f"Les identifiants ont été envoyés avec succès à l'adresse {email}.")
    except Exception as e:
        # Enregistrer dans un fichier pour référence
        import os
        from django.conf import settings
        try:
            log_dir = os.path.join(settings.BASE_DIR, 'sent_emails')
            os.makedirs(log_dir, exist_ok=True)
            with open(os.path.join(log_dir, f"eleve_{eleve.user.username}.txt"), 'w', encoding='utf-8') as f:
                f.write(f"Identifiants pour {eleve.prenom} {eleve.nom}\n")
                f.write(f"Email: {email}\n")
                f.write(f"Identifiant: {eleve.user.username}\n")
                f.write(f"Mot de passe: {password}\n")
            messages.warning(request, f"L'email n'a pas pu être envoyé à {email}, mais les identifiants sont sauvegardés.")
        except Exception as save_error:
            messages.error(request, f"Erreur lors de l'envoi de l'email : {str(e)}")
    
    return redirect('detail_eleve', eleve_id=eleve.id)


# Vue pour envoyer les identifiants d'un professeur par email
@login_required
def envoyer_identifiants_professeur(request, professeur_id):
    from django.conf import settings
    professeur = get_object_or_404(Professeur, id=professeur_id)
    
    # Vérifier si le professeur a un compte utilisateur associé
    if not professeur.user:
        messages.error(request, "Ce professeur n'a pas de compte utilisateur associé.")
        return redirect('detail_professeur', professeur_id=professeur.id)
    
    # Vérifier si l'email est fourni
    email = request.POST.get('email', '')
    if not email:
        messages.error(request, "Veuillez fournir une adresse email valide.")
        return redirect('detail_professeur', professeur_id=professeur.id)
    
    # Générer un nouveau mot de passe
    password = generate_password(professeur.nom)
    professeur.user.set_password(password)
    professeur.user.save()
    
    # Envoyer l'email avec les identifiants
    subject = "Vos identifiants de connexion - École Al Markaz"
    message = f"""Bonjour {professeur.nom},

Voici vos identifiants de connexion pour votre compte École Al Markaz :

Identifiant: {professeur.user.username}
Mot de passe: {password}

Veuillez conserver ces informations en lieu sûr.

Cordialement,
L'équipe de gestion de l'École Al Markaz
    """
    
    # Stocker le mot de passe dans le modèle pour l'afficher plus tard
    professeur.mot_de_passe_en_clair = password
    professeur.save()
    
    # Version HTML du message
    html_message = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
            }}
            .container {{
                max-width: 600px;
                margin: 0 auto;
                padding: 20px;
                border: 1px solid #ddd;
                border-radius: 5px;
            }}
            .header {{
                background-color: #4CAF50;
                color: white;
                padding: 10px;
                text-align: center;
                border-radius: 5px 5px 0 0;
            }}
            .content {{
                padding: 20px;
            }}
            .credentials {{
                background-color: #f9f9f9;
                border: 1px solid #ddd;
                padding: 15px;
                margin: 20px 0;
                border-radius: 5px;
            }}
            .footer {{
                text-align: center;
                margin-top: 20px;
                font-size: 12px;
                color: #777;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>École Al Markaz - Vos identifiants</h2>
            </div>
            <div class="content">
                <p>Bonjour {professeur.nom},</p>
                <p>Voici vos identifiants de connexion pour votre compte École Al Markaz :</p>
                <div class="credentials">
                    <p><strong>Identifiant:</strong> {professeur.user.username}</p>
                    <p><strong>Mot de passe:</strong> {password}</p>
                </div>
                <p>Veuillez conserver ces informations en lieu sûr.</p>
                <p>Cordialement,<br>L'équipe de gestion de l'École Al Markaz</p>
            </div>
            <div class="footer">
                <p>Ce message est automatique, merci de ne pas y répondre.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    sender_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [email]
    
    # Afficher les identifiants directement dans l'interface
    messages.info(request, f"Identifiant: {professeur.user.username}")
    messages.info(request, f"Mot de passe: {password}")
    
    try:
        # Utilisation de EmailMultiAlternatives pour envoyer à la fois du texte et du HTML
        from django.core.mail import EmailMultiAlternatives
        
        # Création de l'email avec le texte simple
        email_message = EmailMultiAlternatives(subject, message, sender_email, recipient_list)
        
        # Ajout du contenu HTML comme alternative
        email_message.attach_alternative(html_message, "text/html")
        
        # Envoi de l'email
        email_message.send()
        
        messages.success(request, f"Les identifiants ont été envoyés avec succès à l'adresse {email}.")
    except Exception as e:
        # Enregistrer dans un fichier pour référence
        import os
        from django.conf import settings
        try:
            log_dir = os.path.join(settings.BASE_DIR, 'sent_emails')
            os.makedirs(log_dir, exist_ok=True)
            with open(os.path.join(log_dir, f"prof_{professeur.user.username}.txt"), 'w', encoding='utf-8') as f:
                f.write(f"Identifiants pour {professeur.nom}\n")
                f.write(f"Email: {email}\n")
                f.write(f"Identifiant: {professeur.user.username}\n")
                f.write(f"Mot de passe: {password}\n")
            messages.warning(request, f"L'email n'a pas pu être envoyé à {email}, mais les identifiants sont sauvegardés.")
        except Exception as save_error:
            messages.error(request, f"Erreur lors de l'envoi de l'email : {str(e)}")
    
    return redirect('detail_professeur', professeur_id=professeur.id)


# Vue pour exporter les élèves filtrés en Excel
@login_required
def export_eleves_excel(request):
    # Récupérer les filtres
    classe_id = request.GET.get('classe')
    creneau_id = request.GET.get('creneau')
    
    # Filtrer les élèves
    eleves = Eleve.objects.all()
    if classe_id:
        eleves = eleves.filter(classes__id=classe_id)
    if creneau_id:
        eleves = eleves.filter(creneau_id=creneau_id)
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Élèves"
    
    # Définir les en-têtes
    headers = ['ID', 'Nom', 'Prénom', 'Classe', 'Créneau', 'Date de naissance', 'Téléphone', 'Email', 'Adresse']
    ws.append(headers)
    
    # Appliquer le style aux en-têtes
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    # Ajouter les données
    for eleve in eleves:
        ws.append([
            eleve.id,
            eleve.nom,
            eleve.prenom,
            ", ".join([classe.nom for classe in eleve.classes.all()]) if eleve.classes.exists() else "",
            ", ".join([classe.creneau.nom for classe in eleve.classes.all() if classe.creneau]) if eleve.classes.exists() else "",
            eleve.date_naissance if eleve.date_naissance else "",
            eleve.telephone or "",
            eleve.email or "",
            eleve.adresse or ""
        ])
    
    # Ajuster les largeurs de colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Sauvegarder et renvoyer le fichier
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="eleves-{datetime.datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

# Vue pour importer des élèves depuis un fichier Excel
@login_required
def import_eleves_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Récupérer les en-têtes
            headers = [str(cell.value).strip() for cell in ws[1]]
            required_headers = ['Nom', 'Prénom']
            
            # Vérifier que les colonnes obligatoires sont présentes
            for header in required_headers:
                if header not in headers:
                    messages.error(request, f'La colonne {header} est obligatoire dans le fichier Excel.')
                    return redirect('liste_eleves')
            
            # Créer les élèves
            eleves_created = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                
                # Vérifier les données obligatoires
                if not row_data.get('Nom') or not row_data.get('Prénom'):
                    continue
                
                # Récupérer la classe et le créneau s'ils existent
                classe = None
                creneau = None
                
                if 'Classe' in headers and row_data.get('Classe'):
                    classe_nom = row_data.get('Classe')
                    classe, created = Classe.objects.get_or_create(nom=classe_nom)
                
                if 'Créneau' in headers and row_data.get('Créneau'):
                    creneau_nom = row_data.get('Créneau')
                    # Formater l'heure début et fin depuis le nom du créneau (format supposé: "HH:MM-HH:MM")
                    creneau_parts = creneau_nom.split('-')
                    if len(creneau_parts) == 2:
                        heure_debut = creneau_parts[0].strip()
                        heure_fin = creneau_parts[1].strip()
                        creneau, created = Creneau.objects.get_or_create(
                            nom=creneau_nom,
                            defaults={
                                'heure_debut': heure_debut,
                                'heure_fin': heure_fin
                            }
                        )
                
                nom = row_data.get('Nom')
                prenom = row_data.get('Prénom')
                
                # Créer un utilisateur pour cet élève
                username = generate_username('eleve', f"{nom}{prenom}")
                password = generate_password(f"{nom} {prenom}")
                
                # Vérifier si le nom d'utilisateur existe déjà
                while User.objects.filter(username=username).exists():
                    username = f"{username}{random.randint(1, 999)}"
                
                # Créer l'utilisateur
                user = User.objects.create_user(username=username, password=password)
                
                # Créer l'élève et associer l'utilisateur
                Eleve.objects.create(
                    nom=nom,
                    prenom=prenom,
                    classe=classe,
                    creneau=creneau,
                    date_naissance=row_data.get('Date de naissance'),
                    telephone=row_data.get('Téléphone', ''),
                    email=row_data.get('Email', ''),
                    adresse=row_data.get('Adresse', ''),
                    user=user
                )
                
                # Stocker les informations d'identification pour l'affichage ultérieur
                if not hasattr(request, 'imported_credentials'):
                    request.imported_credentials = []
                request.imported_credentials.append({
                    'nom': f"{nom} {prenom}",
                    'username': username,
                    'password': password
                })
                eleves_created += 1
            
            # Créer un message pour indiquer le nombre d'élèves importés
            success_msg = f'{eleves_created} élèves ont été importés avec succès!'
            
            # Ajouter les informations d'identification
            if hasattr(request, 'imported_credentials') and request.imported_credentials:
                success_msg += '<br><br><strong>Identifiants de connexion créés:</strong><br>'
                success_msg += '<div style="max-height: 200px; overflow-y: auto; margin-top: 10px;">'
                for cred in request.imported_credentials:
                    success_msg += f"<div>- {cred['nom']}: Identifiant <strong>{cred['username']}</strong> | Mot de passe <strong>{cred['password']}</strong></div>"
                success_msg += '</div>'
                success_msg += '<br><div class="alert alert-warning">Veuillez noter ces identifiants car ils ne seront plus affichés.</div>'
            
            messages.success(request, mark_safe(success_msg))
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'importation: {str(e)}')
    
    return redirect('liste_eleves')


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Eleve, PaiementHistorique
from decimal import Decimal

# Fonction utilitaire pour recalculer le montant total d'un paiement
def recalculer_montant_paiement(paiement):
    """Recalcule le montant total d'un paiement à partir de ses historiques"""
    total = Decimal('0.00')
    for historique in paiement.historiques.all():
        total += historique.montant
    paiement.montant = total
    paiement.save()
    return total

# Vue pour modifier un historique de paiement
@login_required
def modifier_historique_paiement(request, historique_id):
    historique = get_object_or_404(PaiementHistorique, id=historique_id)
    paiement = historique.paiement
    
    if request.method == 'POST':
        from decimal import Decimal
        
        # Récupérer les valeurs directement en Decimal
        nouveau_montant = Decimal(request.POST.get('montant', '0'))
        date = request.POST.get('date')
        methode = request.POST.get('methode')
        commentaire = request.POST.get('commentaire', '')
        
        # Mettre à jour l'historique
        historique.montant = nouveau_montant
        historique.date = date
        historique.methode = methode
        historique.commentaire = commentaire
        historique.save()
        
        # Recalculer le montant total du paiement à partir de tous les historiques
        recalculer_montant_paiement(paiement)
        
        messages.success(request, 'L\'historique de paiement a été modifié avec succès!')
        return redirect('detail_paiement', paiement_id=paiement.id)
    
    context = {
        'historique': historique,
        'paiement': paiement
    }
    return render(request, 'ecole_app/paiements/modifier_historique.html', context)

# Vue pour supprimer un historique de paiement
@login_required
def supprimer_historique_paiement(request, historique_id):
    historique = get_object_or_404(PaiementHistorique, id=historique_id)
    paiement = historique.paiement
    
    if request.method == 'POST':
        # Sauvegarder une référence au paiement avant de supprimer l'historique
        paiement_ref = paiement
        
        # Supprimer l'historique
        historique.delete()
        
        # Recalculer le montant total du paiement à partir des historiques restants
        recalculer_montant_paiement(paiement_ref)
        
        messages.success(request, 'L\'historique de paiement a été supprimé avec succès!')
        return redirect('detail_paiement', paiement_id=paiement.id)
    
    return redirect('detail_paiement', paiement_id=paiement.id)

@login_required
def get_eleve_classe(request):
    eleve_id = request.GET.get('eleve_id')
    try:
        eleve = Eleve.objects.select_related('classe', 'creneau').get(pk=eleve_id)
        data = {
            'classe_id': eleve.classe.id if eleve.classe else None,
            'classe_nom': eleve.classe.nom if eleve.classe else "",
            'creneau_id': eleve.creneau.id if eleve.creneau else None,
            'creneau_nom': eleve.creneau.nom if eleve.creneau else "",
        }
        return JsonResponse(data)
    except Eleve.DoesNotExist:
        return JsonResponse({'error': "Élève introuvable"}, status=404)

# Vue pour envoyer les identifiants à tous les élèves d'un coup
@login_required
def envoyer_tous_identifiants(request):
    if request.method != 'POST':
        messages.error(request, "Méthode non autorisée.")
        return redirect('liste_eleves')
    
    # Vérifier si la confirmation est cochée
    if not request.POST.get('confirm'):
        messages.error(request, "Veuillez confirmer l'envoi des identifiants.")
        return redirect('liste_eleves')
    
    # Récupérer les filtres actuels depuis la session
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante.")
        return redirect('selection_composante')
    
    # Construction de la requête de base (même logique que dans liste_eleves)
    query = Q(composante_id=composante_id) & Q(archive=False)
    
    # Récupérer les filtres de la session ou des paramètres
    classe_id = request.session.get('classe_filter')
    creneau_id = request.session.get('creneau_filter')
    
    # Ajouter les filtres si nécessaire
    if classe_id:
        query &= Q(classes__id=classe_id)
    if creneau_id:
        query &= Q(creneau__id=creneau_id)
    
    # Récupérer tous les élèves correspondant aux critères
    eleves = Eleve.objects.filter(query).distinct()
    
    # Compteurs pour les statistiques
    total_eleves = eleves.count()
    eleves_avec_email = 0
    eleves_sans_email = 0
    eleves_sans_compte = 0
    emails_envoyes = 0
    erreurs = 0
    
    # Traiter chaque élève
    for eleve in eleves:
        # Vérifier si l'élève a un compte utilisateur et un email
        if not eleve.user:
            eleves_sans_compte += 1
            continue
        
        if not eleve.email:
            eleves_sans_email += 1
            continue
        
        eleves_avec_email += 1
        
        try:
            # Générer un nouveau mot de passe au format nom.prenom1
            nom = eleve.nom.lower() if eleve.nom else ''
            prenom = eleve.prenom.lower() if eleve.prenom else ''
            
            # Nettoyer les caractères spéciaux et espaces
            import re
            nom = re.sub(r'[^a-z0-9]', '', nom)
            prenom = re.sub(r'[^a-z0-9]', '', prenom)
            
            # Créer le mot de passe au format nom.prenom1
            password = f"{nom}.{prenom}1" if nom and prenom else generer_mot_de_passe()
            
            # Mettre à jour le mot de passe de l'utilisateur
            eleve.user.set_password(password)
            eleve.user.save()
            
            # Mettre à jour le mot de passe en clair pour l'affichage
            eleve.mot_de_passe_en_clair = password
            eleve.save()
            
            # Envoyer l'email avec les identifiants
            subject = "Vos identifiants de connexion - Al Markaz"
            
            # Message HTML avec des couleurs et un design élaboré
            html_message = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                        margin: 0;
                        padding: 0;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                    }}
                    .header {{
                        background: linear-gradient(135deg, #1e7e34, #218838);
                        color: white;
                        padding: 20px;
                        text-align: center;
                        font-size: 24px;
                        font-weight: bold;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: #f9f9f9;
                        padding: 20px;
                        border-left: 1px solid #ddd;
                        border-right: 1px solid #ddd;
                    }}
                    .arabic {{
                        font-family: 'Traditional Arabic', 'Arial', sans-serif;
                        font-size: 24px;
                        text-align: center;
                        margin: 20px 0;
                        color: #006400;
                        direction: rtl;
                    }}
                    .greeting {{
                        margin-bottom: 20px;
                        font-size: 16px;
                    }}
                    .credentials {{
                        background-color: #f0f8ff;
                        border: 1px solid #4CAF50;
                        border-radius: 5px;
                        padding: 15px;
                        margin: 20px 0;
                    }}
                    .credential-item {{
                        margin-bottom: 10px;
                    }}
                    .credential-label {{
                        font-weight: bold;
                        color: #006400;
                    }}
                    .credential-value {{
                        font-family: monospace;
                        background-color: #f5f5f5;
                        padding: 2px 5px;
                        border-radius: 3px;
                    }}
                    .advice {{
                        background-color: #fffacd;
                        border-left: 4px solid #ffd700;
                        padding: 10px 15px;
                        margin: 20px 0;
                    }}
                    .quote {{
                        font-style: italic;
                        background-color: #e8f5e9;
                        padding: 15px;
                        border-radius: 5px;
                        margin: 20px 0;
                        text-align: center;
                    }}
                    .footer {{
                        background-color: #f5f5f5;
                        padding: 10px 20px;
                        text-align: center;
                        font-size: 12px;
                        color: #666;
                        border-radius: 0 0 5px 5px;
                        border: 1px solid #ddd;
                    }}
                    .mosque-icon {{
                        font-size: 24px;
                        text-align: center;
                        margin: 10px 0;
                    }}
                    .signature {{
                        margin-top: 20px;
                        font-weight: bold;
                    }}
                </style>
            </head>
            <body>
            <div class="container">
                <div class="header">
                    Al Markaz
                </div>
                <div class="content">
                    <div class="arabic">بِسْمِ اللَّهِ الرَّحْمَـٰنِ الرَّحِيمِ</div>
                    
                    <div class="greeting">
                        Assalamu alaykum wa rahmatullahi wa barakatuh <strong>{eleve.prenom} {eleve.nom}</strong>,
                    </div>
                    
                    <p>Nous vous transmettons vos identifiants de connexion :</p>
                    
                    <div class="credentials">
                        <div class="credential-item">
                            <span class="credential-label">📱 Identifiant</span>: 
                            <span class="credential-value">{eleve.user.username}</span>
                        </div>
                        <div class="credential-item">
                            <span class="credential-label">🔐 Mot de passe</span>: 
                            <span class="credential-value">{password}</span>
                        </div>
                    </div>
                    
                    <div class="advice">
                        <p>🌙 <strong>Conseil</strong> : Gardez ces informations confidentielles et changez votre mot de passe régulièrement.</p>
                    </div>
                    
                    <div class="mosque-icon">🕋</div>
                    
                    <div class="quote">
                        "Et rappelle-toi ton Seigneur dans ton âme avec humilité et crainte, et à voix basse, le matin et le soir, et ne sois pas du nombre des négligents."
                        <br><strong>Coran 7:205</strong>
                    </div>
                    
                    <div class="signature">
                        Barakallahu fikum,<br>
                        L'équipe Al Markaz
                    </div>
                </div>
                <div class="footer">
                    &copy; {datetime.datetime.now().year} Al Markaz - Tous droits réservés
                </div>
            </div>
            </body>
            </html>
            """
            
            # Message texte simple pour les clients qui ne supportent pas le HTML
            message = f"""بِسْمِ اللَّهِ الرَّحْمَـٰنِ الرَّحِيمِ

Assalamu alaykum wa rahmatullahi wa barakatuh {eleve.prenom} {eleve.nom},

Nous vous transmettons vos identifiants de connexion :

📱 Identifiant : {eleve.user.username}
🔐 Mot de passe : {password}

🌙 Conseil : Gardez ces informations confidentielles et changez votre mot de passe régulièrement.

📖 "Et rappelle-toi ton Seigneur dans ton âme avec humilité et crainte, et à voix basse, le matin et le soir, et ne sois pas du nombre des négligents." - Coran 7:205

Barakallahu fikum,
L'équipe Al Markaz
"""
            
            sender_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [eleve.email]
            
            # Envoyer l'email avec les deux versions (HTML et texte)
            from django.core.mail import EmailMultiAlternatives
            email_message = EmailMultiAlternatives(subject, message, sender_email, recipient_list)
            email_message.attach_alternative(html_message, "text/html")
            email_message.send()
            
            emails_envoyes += 1
        except Exception as e:
            erreurs += 1
            print(f"Erreur pour l'élève {eleve.id} - {eleve.nom} {eleve.prenom}: {str(e)}")
    
    # Afficher un résumé des opérations
    if emails_envoyes > 0:
        messages.success(request, f"{emails_envoyes} email(s) envoyé(s) avec succès.")
    
    if eleves_sans_email > 0:
        messages.warning(request, f"{eleves_sans_email} élève(s) sans adresse email.")
    
    if eleves_sans_compte > 0:
        messages.warning(request, f"{eleves_sans_compte} élève(s) sans compte utilisateur.")
    
    if erreurs > 0:
        messages.error(request, f"{erreurs} erreur(s) lors de l'envoi des emails.")
    
    return redirect('liste_eleves')

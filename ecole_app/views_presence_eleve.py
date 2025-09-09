from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q
from .models import Eleve, PresenceEleve, Creneau, AnneeScolaire, Classe
import datetime
from django.utils import timezone
from .utils import render_to_pdf
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Fonction pour calculer les statistiques de présence
def calculate_attendance_stats(date, classe, composante_id):
    # Récupérer les élèves de la classe qui ne sont pas archivés
    eleves = Eleve.objects.filter(classes=classe, composante_id=composante_id, archive=False)
    
    # Récupérer les présences pour la date et la classe
    presences = PresenceEleve.objects.filter(
        date=date,
        classe=classe,
        composante_id=composante_id
    )
    
    # Initialiser les statistiques
    stats = {
        'total_eleves': eleves.count(),
        'presents': 0,
        'absents_justifies': 0,
        'absents_non_justifies': 0
    }
    
    # Compter les présences
    for presence in presences:
        if presence.present:
            stats['presents'] += 1
        elif presence.justifie:
            stats['absents_justifies'] += 1
        else:
            stats['absents_non_justifies'] += 1
    
    return stats

@login_required
def gestion_presence_eleve(request):
    """Vue pour gérer les présences des élèves"""
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder à la gestion des présences.")
        return redirect('selection_composante')
    
    # Vérifier si l'utilisateur est un professeur
    est_professeur = hasattr(request.user, 'professeur')
    
    # Récupérer les élèves de la composante sélectionnée qui ne sont pas archivés
    tous_les_eleves = Eleve.objects.filter(composante_id=composante_id, archive=False).order_by('nom')
    
    # Récupérer la date sélectionnée ou utiliser la date du jour
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Récupérer toutes les classes de la composante pour le filtre
    if est_professeur:
        # Si c'est un professeur, ne montrer que ses classes
        toutes_les_classes = Classe.objects.filter(
            composante_id=composante_id, 
            professeur=request.user.professeur
        ).order_by('nom')
    else:
        # Si c'est un admin, montrer toutes les classes de la composante
        toutes_les_classes = Classe.objects.filter(composante_id=composante_id).order_by('nom')
    
    # Récupérer la classe sélectionnée
    classe_id = request.GET.get('classe')
    selected_classe = None
    if classe_id:
        # Nettoyer l'ID de classe pour s'assurer qu'il ne contient que des chiffres
        # Cela évite les erreurs si l'URL contient des paramètres supplémentaires
        if classe_id and '?' in classe_id:
            classe_id = classe_id.split('?')[0]
            
        try:
            # Convertir en entier pour s'assurer que c'est un ID valide
            classe_id = int(classe_id)
            
            # Pour les professeurs, vérifier que la classe leur appartient
            if est_professeur:
                selected_classe = Classe.objects.get(
                    id=classe_id, 
                    professeur=request.user.professeur
                )
            else:
                selected_classe = Classe.objects.get(id=classe_id)
                
            # Si une classe est sélectionnée, filtrer les élèves par cette classe
            tous_les_eleves = tous_les_eleves.filter(classes=selected_classe)
        except (Classe.DoesNotExist, ValueError):
            # Gérer à la fois les erreurs de classe inexistante et de format d'ID invalide
            pass
    
    # Traitement du formulaire POST pour enregistrer une présence
    if request.method == 'POST':
        eleve_id = request.POST.get('eleve_id')
        date = request.POST.get('date')
        classe_id = request.POST.get('classe_id')
        status = request.POST.get('status')
        commentaire = request.POST.get('commentaire', '')
        
        # Déterminer les valeurs de présence et justification selon le statut
        present = False
        justifie = False
        if status == 'present':
            present = True
            justifie = False
        elif status == 'absent_justifie':
            present = False
            justifie = True
        elif status == 'absent':
            present = False
            justifie = False
        
        # Préparer la réponse
        response_data = {
            'success': False,
            'message': '',
            'eleve_id': eleve_id,
            'status': status
        }
        
        if eleve_id and date and classe_id:
            try:
                # Vérifier si l'utilisateur est un professeur
                est_professeur = hasattr(request.user, 'professeur')
                
                # Récupérer l'élève
                eleve = Eleve.objects.get(id=eleve_id)
                
                # Récupérer la classe avec vérification des permissions
                if est_professeur:
                    # Pour un professeur, vérifier que la classe lui appartient
                    try:
                        classe = Classe.objects.get(
                            id=classe_id,
                            professeur=request.user.professeur
                        )
                    except Classe.DoesNotExist:
                        response_data['message'] = "Vous n'avez pas accès à cette classe."
                        return JsonResponse(response_data)
                else:
                    # Pour un admin, pas de restriction
                    classe = Classe.objects.get(id=classe_id)
                
                # Vérifier que l'élève appartient bien à la classe
                if not eleve.classes.filter(id=classe.id).exists():
                    response_data['message'] = "Cet élève n'appartient pas à la classe sélectionnée."
                    return JsonResponse(response_data)
                
                date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                
                # Vérifier si une présence existe déjà pour cet élève, cette date et cette classe
                presence, created = PresenceEleve.objects.get_or_create(
                    eleve=eleve,
                    date=date_obj,
                    classe=classe,
                    defaults={
                        'present': present,
                        'justifie': justifie,
                        'commentaire': commentaire,
                        'composante_id': composante_id
                    }
                )
                
                # Si la présence existait déjà, mettre à jour ses valeurs
                if not created:
                    presence.present = present
                    presence.justifie = justifie
                    presence.commentaire = commentaire
                    presence.save()
                
                if created:
                    message = f'Présence ajoutée pour {eleve}.'
                else:
                    message = f'Présence mise à jour pour {eleve}.'
                messages.success(request, message)
                
                response_data['success'] = True
                response_data['message'] = message
                
                # Recalculer les statistiques pour la mise à jour en temps réel
                if classe_id:
                    try:
                        classe_obj = Classe.objects.get(id=classe_id)
                        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                        stats = calculate_attendance_stats(date_obj, classe_obj, composante_id)
                        response_data['stats'] = stats
                    except Exception as e:
                        print(f"Erreur lors du calcul des statistiques: {str(e)}")
                        # Ne pas bloquer la réponse si le calcul des stats échoue
                
            except (Eleve.DoesNotExist, Classe.DoesNotExist, ValueError) as e:
                error_message = f'Erreur lors de l\'enregistrement de la présence: {str(e)}'
                messages.error(request, error_message)
                response_data['message'] = error_message
        else:
            error_message = 'Données manquantes pour enregistrer la présence (Eleve, Date et Classe requis).'
            messages.error(request, error_message)
            response_data['message'] = error_message
        
        # Si c'est une requête AJAX, renvoyer une réponse JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse(response_data)
        
        # Sinon, rediriger vers la même page avec les mêmes paramètres
        redirect_url = f"?date={date}"
        if classe_id:
            redirect_url += f"&classe={classe_id}"
        return redirect(f"{request.path}{redirect_url}")

    
    # Récupérer les présences existantes pour la date et la classe sélectionnées
    presences_existantes = {}
    if selected_classe:
        presences = PresenceEleve.objects.filter(
            date=selected_date,
            classe=selected_classe,
            composante_id=composante_id
        )
        for presence in presences:
            presences_existantes[presence.eleve.id] = presence
    
    # Construire une liste de tuples (eleve, presence)
    eleves_with_presence = [(eleve, presences_existantes.get(eleve.id, None)) for eleve in tous_les_eleves]
    
    # Statistiques de présence
    stats = {
        'total_eleves': tous_les_eleves.count(),
        'presents': 0,
        'absents_justifies': 0,
        'absents_non_justifies': 0
    }
    
    # Calculer les statistiques si une classe est sélectionnée
    if selected_classe:
        for _, presence in eleves_with_presence:
            if presence:
                if presence.present:
                    stats['presents'] += 1
                elif presence.justifie:
                    stats['absents_justifies'] += 1
                else:
                    stats['absents_non_justifies'] += 1
    
    return render(request, 'ecole_app/eleves/gestion_presence.html', {
        'eleves_with_presence': eleves_with_presence,
        'classes': toutes_les_classes,
        'selected_date': selected_date,
        'selected_classe': selected_classe,
        'stats': stats
    })

@login_required
def rapport_presence_eleve(request):
    """Vue pour afficher un rapport des présences des élèves"""
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder au rapport des présences.")
        return redirect('selection_composante')
    
    # Vérifier si l'utilisateur est un professeur
    is_professeur = False
    professeur = None
    try:
        professeur = request.user.professeur
        is_professeur = True
    except:
        # L'utilisateur n'est pas un professeur (probablement un admin)
        pass
    
    # Si c'est un professeur, filtrer les classes par professeur
    if is_professeur:
        # Récupérer les classes du professeur dans la composante sélectionnée
        classes = Classe.objects.filter(
            composante_id=composante_id,
            professeur=professeur
        ).order_by('nom')
        
        # Récupérer les élèves des classes du professeur
        eleves = Eleve.objects.filter(
            composante_id=composante_id, 
            archive=False,
            classes__in=classes
        ).distinct().order_by('nom')
    else:
        # Pour les administrateurs, afficher tous les élèves et classes de la composante
        eleves = Eleve.objects.filter(composante_id=composante_id, archive=False).order_by('nom')
        classes = Classe.objects.filter(composante_id=composante_id).order_by('nom')
    
    # Filtrage par élève
    eleve_id = request.GET.get('eleve')
    selected_eleve = None
    if eleve_id:
        try:
            # Si c'est un professeur, vérifier que l'élève est dans une de ses classes
            if is_professeur:
                selected_eleve = Eleve.objects.get(id=eleve_id, classes__in=classes)
            else:
                selected_eleve = Eleve.objects.get(id=eleve_id)
        except Eleve.DoesNotExist:
            pass
    
    # Filtrage par classe
    classe_id = request.GET.get('classe')
    selected_classe = None
    if classe_id:
        try:
            # Si c'est un professeur, vérifier que la classe lui appartient
            if is_professeur:
                selected_classe = classes.get(id=classe_id)
            else:
                selected_classe = Classe.objects.get(id=classe_id, composante_id=composante_id)
                
            if not selected_eleve:  # Si aucun élève spécifique n'est sélectionné, filtrer par classe
                eleves = eleves.filter(classes=selected_classe)
        except Classe.DoesNotExist:
            pass
    
    # Filtrage par période
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    date_debut = None
    date_fin = None
    
    if date_debut_str:
        try:
            date_debut = datetime.datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_fin_str:
        try:
            date_fin = datetime.datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if not date_debut:
        # Par défaut, le début du mois courant
        today = timezone.now().date()
        date_debut = datetime.date(today.year, today.month, 1)
    
    if not date_fin:
        # Par défaut, aujourd'hui
        date_fin = timezone.now().date()
    
    # Récupérer les présences
    presences_query = PresenceEleve.objects.filter(
        date__gte=date_debut,
        date__lte=date_fin,
        composante_id=composante_id
    )
    
    # Si c'est un professeur, filtrer les présences par ses classes
    if is_professeur:
        presences_query = presences_query.filter(classe__in=classes)
    
    # Appliquer les filtres supplémentaires
    if selected_eleve:
        presences_query = presences_query.filter(eleve=selected_eleve)
    
    if selected_classe and not selected_eleve:
        presences_query = presences_query.filter(classe=selected_classe)
    
    # Trier les résultats
    presences = presences_query.order_by('date', 'classe__nom')
    
    # Statistiques
    stats = {
        'total': presences.count(),
        'presents': presences.filter(present=True).count(),
        'absents': presences.filter(present=False, justifie=False).count(),
        'absents_justifies': presences.filter(present=False, justifie=True).count(),
    }
    
    if stats['total'] > 0:
        stats['taux_presence'] = round((stats['presents'] / stats['total']) * 100, 1)
    else:
        stats['taux_presence'] = 0
    
    context = {
        'eleves': eleves,
        'classes': classes,
        'selected_eleve': selected_eleve,
        'selected_classe': selected_classe,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'presences': presences,
        'stats': stats,
    }
    
    # Si c'est une requête AJAX, renvoyer les statistiques au format JSON
    if request.GET.get('ajax') == '1':
        return JsonResponse({'stats': stats})
    
    return render(request, 'ecole_app/eleves/rapport_presence.html', context)

@login_required
def export_rapport_presence_excel(request):
    """Exporte le rapport de présence en Excel"""
    
    # Récupération de la classe et des données
    composante_id = request.session.get('composante_id')
    if not composante_id:
        messages.warning(request, "Veuillez sélectionner une composante pour accéder au rapport des présences.")
        return redirect('selection_composante')
    
    # Vérifier si l'utilisateur est un professeur
    is_professeur = False
    professeur = None
    try:
        professeur = request.user.professeur
        is_professeur = True
    except:
        # L'utilisateur n'est pas un professeur (probablement un admin)
        pass
    
    # Si c'est un professeur, filtrer les classes par professeur
    if is_professeur:
        # Récupérer les classes du professeur dans la composante sélectionnée
        classes = Classe.objects.filter(
            composante_id=composante_id,
            professeur=professeur
        ).order_by('nom')
        
        # Récupérer les élèves des classes du professeur
        eleves = Eleve.objects.filter(
            composante_id=composante_id, 
            archive=False,
            classes__in=classes
        ).distinct().order_by('nom')
    else:
        # Pour les administrateurs, afficher tous les élèves et classes de la composante
        eleves = Eleve.objects.filter(composante_id=composante_id, archive=False).order_by('nom')
        classes = Classe.objects.filter(composante_id=composante_id).order_by('nom')
    
    # Filtrage par élève
    eleve_id = request.GET.get('eleve')
    selected_eleve = None
    if eleve_id:
        try:
            # Si c'est un professeur, vérifier que l'élève est dans une de ses classes
            if is_professeur:
                selected_eleve = Eleve.objects.get(id=eleve_id, classes__in=classes)
            else:
                selected_eleve = Eleve.objects.get(id=eleve_id)
        except Eleve.DoesNotExist:
            pass
    
    # Filtrage par classe
    classe_id = request.GET.get('classe')
    selected_classe = None
    if classe_id:
        try:
            # Si c'est un professeur, vérifier que la classe lui appartient
            if is_professeur:
                selected_classe = classes.get(id=classe_id)
            else:
                selected_classe = Classe.objects.get(id=classe_id, composante_id=composante_id)
                
            if not selected_eleve:  # Si aucun élève spécifique n'est sélectionné, filtrer par classe
                eleves = eleves.filter(classes=selected_classe)
        except Classe.DoesNotExist:
            pass
    
    # Filtrage par période
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    date_debut = None
    date_fin = None
    
    if date_debut_str:
        try:
            date_debut = datetime.datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_fin_str:
        try:
            date_fin = datetime.datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if not date_debut:
        # Par défaut, le début du mois courant
        today = timezone.now().date()
        date_debut = datetime.date(today.year, today.month, 1)
    
    if not date_fin:
        # Par défaut, aujourd'hui
        date_fin = timezone.now().date()
    
    # Récupérer les présences
    presences_query = PresenceEleve.objects.filter(
        date__gte=date_debut,
        date__lte=date_fin,
        composante_id=composante_id
    )
    
    # Si c'est un professeur, filtrer les présences par ses classes
    if is_professeur:
        presences_query = presences_query.filter(classe__in=classes)
    
    # Appliquer les filtres supplémentaires
    if selected_eleve:
        presences_query = presences_query.filter(eleve=selected_eleve)
        eleves = [selected_eleve]  # Limiter aux élèves sélectionnés
    
    if selected_classe and not selected_eleve:
        presences_query = presences_query.filter(classe=selected_classe)
        eleves = eleves.filter(classes=selected_classe)
    
    # Trier les résultats
    presences = presences_query.order_by('date', 'classe__nom')
    
    # Statistiques globales
    stats = {
        'total': presences.count(),
        'presents': presences.filter(present=True).count(),
        'absents': presences.filter(present=False, justifie=False).count(),
        'absents_justifies': presences.filter(present=False, justifie=True).count(),
    }
    
    if stats['total'] > 0:
        stats['taux_presence'] = round((stats['presents'] / stats['total']) * 100, 1)
    else:
        stats['taux_presence'] = 0
    
    # Préparer les statistiques par élève pour le format Excel
    eleves_stats = {}
    
    for eleve in eleves:
        # Filtrer les présences pour cet élève
        eleve_presences = presences_query.filter(eleve=eleve)
        
        # Calculer les statistiques pour cet élève
        eleve_stats = {
            'presents': eleve_presences.filter(present=True).count(),
            'absents': eleve_presences.filter(present=False, justifie=False).count(),
            'absents_justifies': eleve_presences.filter(present=False, justifie=True).count(),
            'total': eleve_presences.count(),
            'commentaires': []
        }
        
        # Ajouter les commentaires pour l'historique
        for presence in eleve_presences.exclude(commentaire='').exclude(commentaire__isnull=True).order_by('-date'):
            eleve_stats['commentaires'].append(presence)
        
        # Calculer le taux de présence
        if eleve_stats['total'] > 0:
            eleve_stats['taux_presence'] = round((eleve_stats['presents'] / eleve_stats['total']) * 100, 1)
        else:
            eleve_stats['taux_presence'] = 0
        
        eleves_stats[eleve] = eleve_stats
    
    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Présence {selected_classe.nom if selected_classe else 'Toutes les classes'}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        'Nom',
        'Prénom',
        'Classe actuelle',
        'Présences',
        'Absences justifiées',
        'Absences non justifiées',
        'Taux de présence',
        'Historique des commentaires'
    ]
    
    # Ajout des en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Ajout des données
    row_num = 2
    for eleve, stats in eleves_stats.items():
        # Historique des commentaires formaté
        historique = ""
        if stats['commentaires']:
            historique = "\n".join([f"{c.date.strftime('%d/%m/%Y')}: {c.commentaire}" for c in stats['commentaires']])
        
        data_row = [
            eleve.nom,
            eleve.prenom,
            eleve.classes.first().nom if eleve.classes.first() else '',
            stats['presents'],
            stats['absents_justifies'],
            stats['absents'],
            f"{stats['taux_presence']:.1f}%",
            historique
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top")
            cell.border = border
            
            # Ajustement de la hauteur de ligne pour les commentaires
            if col == 8 and historique:
                ws.row_dimensions[row_num].height = len(historique.split('\n')) * 15
        
        row_num += 1
    
    # Ajustement des largeurs de colonnes
    column_widths = [15, 15, 12, 10, 15, 18, 12, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Création de la réponse HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    mois_fr = [
        '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
        'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
    ]
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_presence_{selected_classe.nom if selected_classe else "toutes_les_classes"}_{date_debut.strftime("%Y%m%d")}_au_{date_fin.strftime("%Y%m%d")}.xlsx"'
    
    return response
